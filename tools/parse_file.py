from crewai.tools import BaseTool
import pandas as pd
from pdf2image import convert_from_path
import easyocr
import numpy as np
import re

class ParseFileTool(BaseTool):
    name: str = "Parse File"
    description: str = (
        "Extrai transações financeiras de arquivos PDF e Excel. "
        "Retorna apenas campos essenciais para reconciliação: data, valor líquido, tipo (PDF) e data_venda, valor_bruto, valor_taxa, valor_liquido, status (Excel)."
    )

    def _run(self,extrato_bancario_path:str):
        """Processa PDF e Excel/CSV e retorna transações essenciais"""
        try:
            # Extrai transações do PDF (extrato bancário)
            pdf_data = self._parse_pdf(extrato_bancario_path)

            # Extrai dados do Excel ou CSV (relatório do adquirente)
            # excel_data = self._parse_excel_or_csv(relatorio_adquirente_path)

            # # Valida erros
            if "error" in pdf_data:
                return {"error": f"Erro no PDF: {pdf_data['error']}"}
            # if "error" in excel_data:
            #     return {"error": f"Erro no Excel/CSV: {excel_data['error']}"}

            return {
                "extrato": pdf_data["transactions"],
                # "relatorio_adquirente": excel_data["relatorio"]
            }

        except Exception as e:
            return {"error": f"Erro no processamento: {str(e)}"}

    def _parse_pdf(self, file_path: str):
        """Extrai texto do PDF via OCR e retorna transações essenciais."""
        try:
            print("🔍 Iniciando OCR no PDF...")
            reader = easyocr.Reader(['pt'], gpu=False)
            images = convert_from_path(file_path, dpi=300)
            print(f"🖼️ {len(images)} páginas convertidas para imagem.")

            texto_total = ""
            for i, image in enumerate(images):
                image_np = np.array(image)
                print(f"🖼️ Processando página {i + 1}...")
                result = reader.readtext(image_np, detail=0, paragraph=True)
                texto_total += "\n".join(result) + "\n"
            print("📜 OCR finalizado.")

            if not texto_total.strip():
                return {"error": "Nenhum texto extraído do PDF"}

            transactions = self._extract_transactions_from_text(texto_total)
            return {"transactions": transactions}

        except Exception as e:
            return {"error": f"Erro ao processar PDF: {str(e)}"}

    def _extract_transactions_from_text(self, text: str):
            """Extrai transações do texto usando regex."""
            transactions = []
            pattern = re.compile(
                r'(\d{2}/\d{2}/\d{4})\s+(.+?)\s+(-?[\d\.,]+)\s+(-?[\d\.,]+)'
            )

            for match in pattern.finditer(text):
                try:
                    data = match.group(1)
                    descricao = match.group(2).strip()
                    valor_bruto = float(match.group(3).replace('.', '').replace(',', '.'))
                    valor_liquido = float(match.group(4).replace('.', '').replace(',', '.'))

                    transacao = {
                        "data": data,
                        "descricao": descricao,
                        "valor_bruto": valor_bruto,
                        "valor_liquido": valor_liquido,
                        "tipo": None
                    }

                    if self._validar_transacao(transacao):
                        transactions.append(transacao)
                except Exception as e:
                    print(f"⚠️ Erro ao processar transação: {e}")

            return transactions
        
    def _parse_excel_or_csv(self, file_path: str):
        try:
            import openpyxl

            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active

            # Mapeia colunas pelo cabeçalho
            header_map = {}
            for row in sheet.iter_rows(min_row=1, max_row=50):  # Procura cabeçalho nas primeiras 50 linhas
                for cell in row:
                    if cell.value:
                        cell_value = str(cell.value).strip().lower()
                        if "data da venda" in cell_value:
                            header_map["data_venda"] = cell.column - 1
                        elif "valor bruto" in cell_value:
                            header_map["valor_bruto"] = cell.column - 1
                        elif "valor da taxa" in cell_value:
                            header_map["valor_taxa"] = cell.column - 1
                        elif "valor líquido" in cell_value:
                            header_map["valor_liquido"] = cell.column - 1
                        elif "status" in cell_value:
                            header_map["status"] = cell.column - 1
                            

            # Extrai transações usando os índices mapeados
            relatorio = []
            for row in sheet.iter_rows(min_row=1, max_row=10000):  # Lê até 10000 linhas
                data = {}
                for cell in row:
                    col_idx = cell.column - 1
                    if col_idx in header_map.values():
                        key = next(k for k, v in header_map.items() if v == col_idx)
                        cell_value = str(cell.value).strip() if cell.value else ""

                        # Converte valores monetários
                        if key in ["valor_bruto", "valor_taxa", "valor_liquido"]:
                            if re.search(r'[\d.,]+', cell_value):
                                valor_str = cell_value.replace("R$", "").replace(" ", "")
                                if "." in valor_str and "," in valor_str:
                                    valor = float(valor_str.replace(".", "").replace(",", "."))
                                elif "," in valor_str:
                                    valor = float(valor_str.replace(",", "."))
                                else:
                                    valor = float(valor_str)
                                data[key] = valor
                        elif key == "data_venda" and re.match(r'\d{2}/\d{2}/\d{4}', cell_value):
                            data[key] = cell_value
                        elif key == "status" and cell_value in ["Aprovada", "Cancelada"]:
                            data[key] = cell_value

                
                            

                # Valida transação completa
                if all(k in data for k in ["data_venda", "valor_bruto", "valor_taxa", "valor_liquido", "status"]):
                    relatorio.append(data)

                   

            return {"relatorio": relatorio}

        except Exception as e:
            return {"error": f"Erro ao processar Excel/CSV: {str(e)}"}
    def _validar_transacao(self, transacao: dict) -> bool:
        """Valida campos essenciais da transação"""
        try:
            # Valida data
            data_parts = transacao["data"].split("/")
            if len(data_parts) != 3:
                return False
            
            dia, mes, ano = map(int, data_parts)
            if not (1 <= dia <= 31 and 1 <= mes <= 12 and ano > 1900):
                return False
                
            # Valida valores numéricos
            for campo in ["valor_liquido"]:
                if not isinstance(transacao[campo], (int, float)):
                    return False
                    
            return True
        except Exception:
            return False